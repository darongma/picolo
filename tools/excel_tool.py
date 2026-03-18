"""
Read, create, and edit Excel (.xlsx) files.
"""
import openpyxl
from openpyxl.utils import get_column_letter

# Read tool
tool_spec_read = {
    "name": "excel_read",
    "description": "Read data from an Excel file (.xlsx). Specify sheet and optional cell range. Returns a plain text table representation.",
    "parameters": {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Path to the Excel file."
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name or 0-based index. Default: first sheet."
            },
            "cell_range": {
                "type": "string",
                "description": "Cell range in A1 notation (e.g., 'A1:C10'). If omitted, reads all used cells."
            }
        },
        "required": ["file_path"]
    }
}

def _get_sheet(wb, sheet):
    if sheet is None:
        return wb.active
    if isinstance(sheet, int):
        try:
            return wb.worksheets[sheet]
        except IndexError:
            raise ValueError(f"Sheet index {sheet} out of range")
    else:
        return wb[sheet]

def run_read(file_path: str, sheet=None, cell_range=None) -> str:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = _get_sheet(wb, sheet)
        
        if cell_range:
            rows = ws[cell_range]
        else:
            rows = ws.iter_rows(values_only=True)
        
        lines = []
        for row in rows:
            line = "\t".join(str(cell) if cell is not None else "" for cell in row)
            lines.append(line)
        
        wb.close()
        return "\n".join(lines) if lines else "(sheet is empty)"
    except Exception as e:
        return f"Error: {e}"

# Create tool
tool_spec_create = {
    "name": "excel_create",
    "description": "Create a new Excel file (.xlsx) with the given 2D data. Writes to a single sheet named 'Sheet1'.",
    "parameters": {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Path where the .xlsx file will be saved."
            },
            "data": {
                "type": "array",
                "items": {
                    "type": "array",
                    "items": {"type": "string"}
                },
                "description": "2D array (list of rows) containing the cell values. Each row is a list of strings/numbers."
            }
        },
        "required": ["file_path", "data"]
    }
}

def run_create(file_path: str, data: list) -> str:
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        for r_idx, row in enumerate(data, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(file_path)
        wb.close()
        return f"Created Excel file with {len(data)} rows at {file_path}"
    except Exception as e:
        return f"Error: {e}"

# Edit tool
tool_spec_edit = {
    "name": "excel_edit",
    "description": "Edit an existing Excel file. Write a 2D data array starting at a specified cell (default A1). Overwrites existing cells in the range.",
    "parameters": {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Path to the .xlsx file to edit."
            },
            "data": {
                "type": "array",
                "items": {
                    "type": "array",
                    "items": {"type": "string"}
                },
                "description": "2D array (list of rows) to write."
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name or index (0-based). Default: active/first sheet."
            },
            "start_cell": {
                "type": "string",
                "description": "Top-left cell to start writing (A1 notation). Default: 'A1'."
            }
        },
        "required": ["file_path", "data"]
    }
}

def run_edit(file_path: str, data: list, sheet=None, start_cell="A1") -> str:
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet is None:
            ws = wb.active
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        
        # Determine start row and column
        start_col_letter = ''.join(filter(str.isalpha, start_cell.upper()))
        start_row_str = ''.join(filter(str.isdigit, start_cell))
        start_row = int(start_row_str) if start_row_str else 1
        start_col = openpyxl.utils.column_index_from_string(start_col_letter) if start_col_letter else 1
        
        for r_idx, row in enumerate(data, start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(file_path)
        wb.close()
        return f"Wrote {len(data)} rows starting at {start_cell} in {file_path}"
    except Exception as e:
        return f"Error: {e}"

tool_specs = [tool_spec_read, tool_spec_create, tool_spec_edit]

tools = {
    "excel_read": run_read,
    "excel_create": run_create,
    "excel_edit": run_edit
}
